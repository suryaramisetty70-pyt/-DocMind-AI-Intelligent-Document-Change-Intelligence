"""
DocMind AI - FastAPI Application v3.0
Properly parses DOCX, PDF, Excel, and TXT files
"""

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import uuid
from pathlib import Path
import tempfile
from typing import Optional

app = FastAPI(title="DocMind AI API", version="3.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

TEMP_DIR = Path(tempfile.gettempdir()) / "docmind_ai"
TEMP_DIR.mkdir(parents=True, exist_ok=True)


def extract_text_from_file(file_path: Path, file_ext: str) -> str:
    """Extract text content from various file types"""
    
    # TXT file - simple text
    if file_ext in ['.txt', '.text']:
        try:
            return file_path.read_text(encoding='utf-8', errors='ignore')
        except:
            return file_path.read_text(encoding='latin-1', errors='ignore')
    
    # DOCX file - Word document
    elif file_ext in ['.docx', '.doc']:
        try:
            from docx import Document
            doc = Document(file_path)
            return '\n\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
        except ImportError:
            return "ERROR: python-docx not installed. Run: pip install python-docx"
        except Exception as e:
            return f"ERROR parsing DOCX: {str(e)}"
    
    # PDF file
    elif file_ext == '.pdf':
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                text = []
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text.append(page_text)
                return '\n\n'.join(text)
        except ImportError:
            return "ERROR: pdfplumber not installed. Run: pip install pdfplumber"
        except Exception as e:
            return f"ERROR parsing PDF: {str(e)}"
    
    # Excel file
    elif file_ext in ['.xlsx', '.xls']:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            all_sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_text = ' | '.join([str(cell) if cell is not None else '' for cell in row])
                    if row_text.strip():
                        rows.append(row_text)
                if rows:
                    all_sheets.append(f"=== {sheet_name} ===\n" + '\n'.join(rows))
            return '\n\n'.join(all_sheets)
        except ImportError:
            return "ERROR: openpyxl not installed. Run: pip install openpyxl"
        except Exception as e:
            return f"ERROR parsing Excel: {str(e)}"
    
    # Unknown file type - try as text
    else:
        try:
            return file_path.read_text(encoding='utf-8', errors='ignore')
        except:
            return "ERROR: Cannot read this file type"


@app.get("/")
async def root():
    return {"name": "DocMind AI API", "version": "3.0", "status": "running"}


@app.get("/api/v1/health")
async def health():
    return {"status": "healthy", "version": "3.0", "message": "API server is running correctly!"}


@app.post("/api/v1/documents/upload")
async def upload_document(file: UploadFile = File(...)):
    """Upload and parse a document"""
    file_ext = Path(file.filename or "file").suffix.lower()
    temp_path = TEMP_DIR / f"upload_{uuid.uuid4()}{file_ext}"
    
    try:
        content = file.file.read()
        temp_path.write_bytes(content)
        text = extract_text_from_file(temp_path, file_ext)
        return {
            "filename": file.filename,
            "file_type": file_ext,
            "content_length": len(text),
            "preview": text[:500] if len(text) > 500 else text,
            "status": "uploaded"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if temp_path.exists():
            temp_path.unlink()


@app.post("/api/v1/compare")
async def compare_documents(original_file: UploadFile = File(...), modified_file: UploadFile = File(...)):
    """Compare two documents"""
    orig_ext = Path(original_file.filename or "original").suffix.lower()
    mod_ext = Path(modified_file.filename or "modified").suffix.lower()
    
    orig_path = TEMP_DIR / f"orig_{uuid.uuid4()}{orig_ext}"
    mod_path = TEMP_DIR / f"mod_{uuid.uuid4()}{mod_ext}"
    
    try:
        # Save files
        orig_path.write_bytes(original_file.file.read())
        mod_path.write_bytes(modified_file.file.read())
        
        # Extract text from both files
        original_text = extract_text_from_file(orig_path, orig_ext)
        modified_text = extract_text_from_file(mod_path, mod_ext)
        
        # Check for errors
        if original_text.startswith("ERROR:"):
            return {"error": f"Original file: {original_text}"}
        if modified_text.startswith("ERROR:"):
            return {"error": f"Modified file: {modified_text}"}
        
        # Compare texts
        orig_lines = [l for l in original_text.split('\n') if l.strip()]
        mod_lines = [l for l in modified_text.split('\n') if l.strip()]
        
        changes = []
        max_lines = max(len(orig_lines), len(mod_lines))
        
        for i in range(max_lines):
            orig_line = orig_lines[i] if i < len(orig_lines) else ""
            mod_line = mod_lines[i] if i < len(mod_lines) else ""
            
            if orig_line != mod_line:
                changes.append({
                    "change_id": i + 1,
                    "type": "modified" if orig_line and mod_line else ("insertion" if not orig_line else "deletion"),
                    "original_content": orig_line if orig_line else "(empty)",
                    "modified_content": mod_line if mod_line else "(empty)",
                    "line_number": i + 1
                })
        
        # Calculate similarity
        matching_lines = sum(1 for i in range(min(len(orig_lines), len(mod_lines))) if orig_lines[i] == mod_lines[i])
        similarity = (matching_lines / max_lines * 100) if max_lines > 0 else 100
        
        # Count insertions and deletions
        orig_set = set(orig_lines)
        mod_set = set(mod_lines)
        insertions = len(mod_set - orig_set)
        deletions = len(orig_set - mod_set)
        
        return {
            "status": "success",
            "message": f"Compared: {original_file.filename} vs {modified_file.filename}",
            "similarity": round(similarity, 2),
            "total_changes": len(changes),
            "insertions": insertions,
            "deletions": deletions,
            "changes": changes[:50],  # Limit to 50 changes
            "original_preview": original_text[:200],
            "modified_preview": modified_text[:200]
        }
    
    except Exception as e:
        return {"error": f"Comparison failed: {str(e)}"}
    
    finally:
        if orig_path.exists():
            orig_path.unlink()
        if mod_path.exists():
            mod_path.unlink()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
