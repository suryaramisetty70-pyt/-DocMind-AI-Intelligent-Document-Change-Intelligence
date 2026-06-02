"""
DocMind AI - Report Generator
Generate PDF, Excel, and comprehensive reports
"""

from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional
from enum import Enum
from pathlib import Path
import json
from datetime import datetime
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Font = None
    PatternFill = None
    Alignment = None
    Border = None
    Side = None
    OPENPYXL_AVAILABLE = False


class ReportFormat(Enum):
    """Supported report formats"""
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"
    JSON = "json"
    MARKDOWN = "markdown"


class ReportType(Enum):
    """Types of reports"""
    EXECUTIVE = "executive"
    DETAILED = "detailed"
    AUDIT = "audit"
    COMPLIANCE = "compliance"
    FRAUD = "fraud"


@dataclass
class ReportSection:
    """Individual report section"""
    title: str
    content: Any
    order: int
    include: bool = True


class PDFReportGenerator:
    """Generate PDF reports using ReportLab"""
    
    def __init__(self):
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib import colors
            self.reportlab_available = True
            self.letter = letter
            self.A4 = A4
            self.getSampleStyleSheet = getSampleStyleSheet
            self.ParagraphStyle = ParagraphStyle
            self.inch = inch
            self.SimpleDocTemplate = SimpleDocTemplate
            self.Paragraph = Paragraph
            self.Spacer = Spacer
            self.Table = Table
            self.TableStyle = TableStyle
            self.PageBreak = PageBreak
            self.colors = colors
        except ImportError:
            self.reportlab_available = False
    
    def generate(self, sections: List[ReportSection], output_path: str, title: str = "DocMind AI Report") -> bool:
        """Generate PDF report"""
        if not self.reportlab_available:
            raise ImportError("ReportLab not installed. Run: pip install reportlab")
        
        doc = self.SimpleDocTemplate(output_path, pagesize=self.A4)
        styles = self.getSampleStyleSheet()
        
        # Custom styles
        title_style = self.ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        heading_style = self.ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=10,
            textColor=self.colors.HexColor('#1E3A5F')
        )
        
        story = []
        
        # Title
        story.append(self.Paragraph(title, title_style))
        story.append(self.Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(self.Spacer(1, 0.5 * self.inch))
        
        # Sections
        for section in sorted(sections, key=lambda s: s.order):
            if section.include:
                story.append(self.Paragraph(section.title, heading_style))
                story.append(self._render_content(section.content, styles))
                story.append(self.Spacer(1, 0.3 * self.inch))
        
        doc.build(story)
        return True
    
    def _render_content(self, content: Any, styles) -> Any:
        """Render content based on type"""
        if isinstance(content, str):
            return self.Paragraph(content, styles['Normal'])
        elif isinstance(content, list):
            elements = []
            for item in content:
                if isinstance(item, str):
                    elements.append(self.Paragraph(item, styles['Normal']))
                    elements.append(self.Spacer(1, 0.1 * self.inch))
                elif isinstance(item, dict):
                    elements.append(self._render_dict(item, styles))
            return elements
        elif isinstance(content, dict):
            return self._render_dict(content, styles)
        
        return self.Paragraph(str(content), styles['Normal'])
    
    def _render_dict(self, data: Dict, styles) -> Any:
        """Render dictionary as table or paragraphs"""
        if not data:
            return self.Paragraph("No data available", styles['Normal'])
        
        # Check if it's a simple key-value dict
        if all(isinstance(v, (str, int, float)) for v in data.values()):
            # Render as table
            table_data = [[k, str(v)] for k, v in data.items()]
            table = self.Table(table_data, colWidths=[2 * self.inch, 4 * self.inch])
            table.setStyle(self.TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), self.colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), self.colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, self.colors.grey)
            ]))
            return table
        
        # Otherwise render as paragraphs
        elements = []
        for key, value in data.items():
            elements.append(self.Paragraph(f"<b>{key}:</b> {value}", styles['Normal']))
            elements.append(self.Spacer(1, 0.1 * self.inch))
        return elements


class ExcelReportGenerator:
    """Generate Excel reports using openpyxl"""
    
    def __init__(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            self.openpyxl = openpyxl
            self.Font = Font
            self.PatternFill = PatternFill
            self.Alignment = Alignment
            self.Border = Border
            self.Side = Side
            self.get_column_letter = get_column_letter
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
    
    def generate(self, sections: List[ReportSection], output_path: str, title: str = "DocMind AI Report") -> bool:
        """Generate Excel report"""
        if not self.openpyxl_available:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = self.openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Styles
        title_font = self.Font(bold=True, size=16, color="FFFFFF")
        title_fill = self.PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = self.Font(bold=True, size=12)
        header_fill = self.PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        
        # Create sheets from sections
        for section in sorted(sections, key=lambda s: s.order):
            if section.include:
                ws = wb.create_sheet(title=section.title[:31])  # Excel sheet name max 31 chars
                
                # Title row
                ws.merge_cells('A1:E1')
                ws['A1'] = section.title
                ws['A1'].font = title_font
                ws['A1'].fill = title_fill
                ws['A1'].alignment = self.Alignment(horizontal='center')
                
                # Content
                self._populate_sheet(ws, section.content, header_font, header_fill)
        
        wb.save(output_path)
        return True
    
    def _populate_sheet(self, ws, content: Any, header_font: Font, header_fill: PatternFill):
        """Populate worksheet with content"""
        if isinstance(content, dict):
            row = 3
            for key, value in content.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value))
                row += 1
        
        elif isinstance(content, list):
            row = 3
            for idx, item in enumerate(content):
                if isinstance(item, dict):
                    if idx == 0:
                        # Header row
                        for col, key in enumerate(item.keys(), 1):
                            cell = ws.cell(row=row, column=col, value=key)
                            cell.font = header_font
                            cell.fill = header_fill
                        row += 1
                    
                    # Data row
                    for col, value in enumerate(item.values(), 1):
                        ws.cell(row=row, column=col, value=str(value) if value else "")
                    row += 1
                elif isinstance(item, str):
                    ws.cell(row=row, column=1, value=item)
                    row += 1
        
        # Auto-adjust column widths
        for ws in ws.parent.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width


class ReportGenerator:
    """Main report generator coordinating all formats"""
    
    def __init__(self):
        self.pdf_generator = PDFReportGenerator()
        self.excel_generator = ExcelReportGenerator()
    
    def generate_report(
        self,
        comparison_result: Any,
        semantic_result: Any,
        risk_result: Any,
        fraud_result: Any,
        similarity_result: Any,
        output_path: str,
        format: ReportFormat = ReportFormat.PDF,
        report_type: ReportType = ReportType.EXECUTIVE
    ) -> bool:
        """Generate comprehensive report"""
        
        # Build sections
        sections = self._build_sections(
            comparison_result,
            semantic_result,
            risk_result,
            fraud_result,
            similarity_result,
            report_type
        )
        
        # Generate based on format
        if format == ReportFormat.PDF:
            return self.pdf_generator.generate(sections, output_path)
        elif format == ReportFormat.EXCEL:
            return self.excel_generator.generate(sections, output_path)
        elif format == ReportFormat.JSON:
            return self._generate_json(sections, output_path)
        elif format == ReportFormat.MARKDOWN:
            return self._generate_markdown(sections, output_path)
        
        return False
    
    def _build_sections(
        self,
        comparison_result: Any,
        semantic_result: Any,
        risk_result: Any,
        fraud_result: Any,
        similarity_result: Any,
        report_type: ReportType
    ) -> List[ReportSection]:
        """Build report sections based on type"""
        sections = []
        
        # Overview section
        sections.append(ReportSection(
            title="Overview",
            content={
                "Total Changes": len(comparison_result.changes) if comparison_result else 0,
                "Overall Similarity": f"{comparison_result.overall_similarity * 100:.1f}%" if comparison_result else "N/A",
                "Semantic Similarity": f"{semantic_result.overall_semantic_similarity * 100:.1f}%" if semantic_result else "N/A",
                "Processing Time": f"{comparison_result.processing_time:.2f}s" if comparison_result else "N/A"
            },
            order=1
        ))
        
        # Similarity section
        if similarity_result:
            sections.append(ReportSection(
                title="Similarity Analysis",
                content={
                    "Overall Similarity": f"{similarity_result.overall_similarity * 100:.1f}%",
                    "Semantic Similarity": f"{similarity_result.semantic_similarity * 100:.1f}%",
                    "Structural Similarity": f"{similarity_result.structural_similarity * 100:.1f}%",
                    "Lexical Similarity": f"{similarity_result.lexical_similarity * 100:.1f}%"
                },
                order=2
            ))
        
        # Changes section (for detailed reports)
        if report_type in [ReportType.DETAILED, ReportType.AUDIT] and comparison_result:
            changes_data = []
            for change in comparison_result.changes[:100]:  # Limit to 100
                changes_data.append({
                    "Type": change.change_type.value,
                    "Severity": change.severity.value,
                    "Category": change.category.value,
                    "Original": change.original_content[:50] if change.original_content else "",
                    "Modified": change.modified_content[:50] if change.modified_content else ""
                })
            
            sections.append(ReportSection(
                title="Changes Detail",
                content=changes_data,
                order=3
            ))
        
        # Risk section
        if risk_result:
            sections.append(ReportSection(
                title="Risk Analysis",
                content={
                    "Overall Risk Score": f"{risk_result.overall_risk_score * 100:.1f}%",
                    "Risk Level": risk_result.risk_level.value if hasattr(risk_result, 'risk_level') else "Unknown",
                    "Financial Risk": f"{risk_result.financial_risk * 100:.1f}%" if hasattr(risk_result, 'financial_risk') else "N/A",
                    "Legal Risk": f"{risk_result.legal_risk * 100:.1f}%" if hasattr(risk_result, 'legal_risk') else "N/A",
                    "Compliance Risk": f"{risk_result.compliance_risk * 100:.1f}%" if hasattr(risk_result, 'compliance_risk') else "N/A"
                },
                order=4
            ))
        
        # Fraud section
        if fraud_result and report_type in [ReportType.FRAUD, ReportType.EXECUTIVE]:
            sections.append(ReportSection(
                title="Fraud Detection",
                content={
                    "Fraud Score": f"{fraud_result.fraud_score * 100:.1f}%",
                    "Fraud Level": fraud_result.fraud_level.value if hasattr(fraud_result, 'fraud_level') else "Unknown",
                    "Total Indicators": len(fraud_result.indicators),
                    "Critical Findings": len(fraud_result.critical_findings) if hasattr(fraud_result, 'critical_findings') else 0
                },
                order=5
            ))
        
        # Recommendations section
        if risk_result and risk_result.recommendations:
            sections.append(ReportSection(
                title="Recommendations",
                content=risk_result.recommendations[:10],
                order=10
            ))
        
        return sections
    
    def _generate_json(self, sections: List[ReportSection], output_path: str) -> bool:
        """Generate JSON report"""
        data = {
            "generated_at": datetime.now().isoformat(),
            "sections": []
        }
        
        for section in sections:
            data["sections"].append({
                "title": section.title,
                "content": section.content,
                "order": section.order
            })
        
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2, default=str)
        
        return True
    
    def _generate_markdown(self, sections: List[ReportSection], output_path: str) -> bool:
        """Generate Markdown report"""
        md = "# DocMind AI Analysis Report\n\n"
        md += f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        
        for section in sorted(sections, key=lambda s: s.order):
            if section.include:
                md += f"## {section.title}\n\n"
                md += self._markdown_content(section.content)
                md += "\n\n"
        
        with open(output_path, 'w') as f:
            f.write(md)
        
        return True
    
    def _markdown_content(self, content: Any) -> str:
        """Convert content to markdown"""
        if isinstance(content, dict):
            lines = []
            for key, value in content.items():
                lines.append(f"- **{key}:** {value}")
            return "\n".join(lines)
        elif isinstance(content, list):
            if all(isinstance(item, dict) for item in content):
                # Table
                if content:
                    headers = list(content[0].keys())
                    md = "| " + " | ".join(headers) + " |\n"
                    md += "| " + " | ".join(["---"] * len(headers)) + " |\n"
                    for row in content:
                        md += "| " + " | ".join(str(row.get(h, ""))[:50] for h in headers) + " |\n"
                    return md
            else:
                return "\n".join(f"- {item}" for item in content)
        
        return str(content)


class AuditReportGenerator:
    """Generate audit-specific reports"""
    
    def generate_audit_trail(
        self,
        comparison_id: str,
        original_document: str,
        modified_document: str,
        changes: List[Any],
        user_actions: List[Dict],
        output_path: str
    ) -> bool:
        """Generate audit trail report"""
        
        report = f"""# Audit Trail Report

## Document Information

- **Comparison ID:** {comparison_id}
- **Original Document:** {original_document}
- **Modified Document:** {modified_document}
- **Report Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Change Summary

- **Total Changes:** {len(changes)}

### Changes by Type
"""
        
        # Count by type
        type_counts = {}
        for change in changes:
            ct = change.change_type.value
            type_counts[ct] = type_counts.get(ct, 0) + 1
        
        for ct, count in type_counts.items():
            report += f"- {ct}: {count}\n"
        
        report += "\n### Changes by Severity\n"
        
        severity_counts = {}
        for change in changes:
            sev = change.severity.value
            severity_counts[sev] = severity_counts.get(sev, 0) + 1
        
        for sev, count in severity_counts.items():
            report += f"- {sev}: {count}\n"
        
        report += "\n## Detailed Change Log\n\n"
        
        for idx, change in enumerate(changes, 1):
            report += f"""### Change {idx}

- **Type:** {change.change_type.value}
- **Severity:** {change.severity.value}
- **Category:** {change.category.value}
- **Location:** Page {change.location.page or 'N/A'}, Line {change.location.line_start or 'N/A'}

**Original:**
```
{change.original_content or '(empty)'}
```

**Modified:**
```
{change.modified_content or '(empty)'}
```

---
"""
        
        report += "\n## User Actions\n\n"
        
        for action in user_actions:
            report += f"- {action.get('timestamp', 'N/A')}: {action.get('action', 'N/A')} - {action.get('user', 'N/A')}\n"
        
        with open(output_path, 'w') as f:
            f.write(report)
        
        return True