"""
DocMind AI - Excel Parser
Handles XLS and XLSX document parsing with multi-sheet support
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl import load_workbook
import pandas as pd

from .base_parser import (
    BaseDocumentParser,
    DocumentContent,
    DocumentMetadata,
    DocumentElement,
    DocumentParserFactory
)


class ExcelParser(BaseDocumentParser):
    """Parser for Excel documents (XLS/XLSX)"""
    
    SUPPORTED_EXTENSIONS = ["xlsx", "xls", "xlsm"]
    
    def __init__(self):
        super().__init__()
        self.extract_formulas = True
        self.extract_hidden_rows = True
        self.extract_comments = True
    
    def parse(self, file_path: Path) -> DocumentContent:
        """Parse Excel file and extract structured content"""
        self.validate_file(file_path)
        
        metadata = self.extract_metadata(file_path)
        text_content = self.extract_text(file_path)
        elements = self._extract_elements(file_path)
        tables = self._extract_tables(file_path)
        structure = self._analyze_structure(file_path)
        
        # Get sheet info for metadata
        wb = load_workbook(file_path, data_only=True)
        metadata.page_count = len(wb.sheetnames)
        wb.close()
        
        return DocumentContent(
            metadata=metadata,
            text_content=text_content,
            elements=elements,
            tables=tables,
            images=[],
            structure=structure
        )
    
    def extract_text(self, file_path: Path) -> str:
        """Extract text from Excel file"""
        text_parts = []
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"<!-- Sheet: {sheet_name} -->")
                
                for row in sheet.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            # Handle different data types
                            if hasattr(cell.value, 'strftime'):
                                row_values.append(str(cell.value.strftime('%Y-%m-%d %H:%M:%S')))
                            else:
                                row_values.append(str(cell.value))
                    
                    if row_values:
                        text_parts.append(" | ".join(row_values))
                
                text_parts.append("")  # Empty line between sheets
            
            wb.close()
        except Exception as e:
            # Try with pandas for older .xls files
            try:
                df = pd.read_excel(file_path, sheet_name=None)
                for sheet_name, sheet_df in df.items():
                    text_parts.append(f"<!-- Sheet: {sheet_name} -->")
                    text_parts.append(sheet_df.to_string())
                    text_parts.append("")
            except Exception as e2:
                raise ValueError(f"Failed to parse Excel file: {e2}")
        
        return "\n".join(text_parts)
    
    def _extract_elements(self, file_path: Path) -> List[DocumentElement]:
        """Extract structured elements from Excel"""
        elements = []
        
        try:
            wb = load_workbook(file_path)
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                sheet = wb[sheet_name]
                
                # Extract cells with their coordinates
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            element = DocumentElement(
                                element_type="cell",
                                content=str(cell.value),
                                page_number=sheet_idx,
                                position=(cell.column, cell.row, 0, 0),
                                attributes={
                                    "coordinate": cell.coordinate,
                                    "data_type": str(cell.data_type),
                                    "formula": cell.value if self.extract_formulas and isinstance(cell.value, str) and cell.value.startswith('=') else None
                                }
                            )
                            elements.append(element)
                
                # Extract merged cells
                for merged_range in sheet.merged_cells.ranges:
                    element = DocumentElement(
                        element_type="merged_cell",
                        content=f"Merged: {merged_range}",
                        page_number=sheet_idx
                    )
                    elements.append(element)
            
            wb.close()
        except Exception as e:
            pass
        
        return elements
    
    def _extract_tables(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extract tables from Excel"""
        tables = []
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                sheet = wb[sheet_name]
                
                # Check for data range (table-like structure)
                data_rows = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        data_rows.append(list(row))
                
                if data_rows:
                    tables.append({
                        "sheet": sheet_name,
                        "sheet_index": sheet_idx,
                        "data": data_rows,
                        "rows": len(data_rows),
                        "columns": len(data_rows[0]) if data_rows else 0,
                        "has_header": True  # Assume first row is header
                    })
            
            wb.close()
        except Exception as e:
            pass
        
        return tables
    
    def _analyze_structure(self, file_path: Path) -> Dict[str, Any]:
        """Analyze Excel file structure"""
        structure = {
            "sheets": [],
            "total_rows": 0,
            "total_columns": 0,
            "formulas_count": 0,
            "charts_count": 0,
            "images_count": 0
        }
        
        try:
            wb = load_workbook(file_path)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                max_row = sheet.max_row or 0
                max_col = sheet.max_column or 0
                
                # Count formulas
                formulas = 0
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas += 1
                
                sheet_info = {
                    "name": sheet_name,
                    "max_row": max_row,
                    "max_column": max_col,
                    "formulas": formulas
                }
                structure["sheets"].append(sheet_info)
                structure["total_rows"] += max_row
                structure["total_columns"] = max(structure["total_columns"], max_col)
                structure["formulas_count"] += formulas
            
            structure["charts_count"] = sum(len(wb[s]._charts) for s in wb.sheetnames)
            
            wb.close()
        except Exception as e:
            pass
        
        return structure
    
    def get_sheet_names(self, file_path: Path) -> List[str]:
        """Get all sheet names from Excel file"""
        wb = load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    
    def get_cell_value(self, file_path: Path, sheet_name: str, cell_address: str) -> Any:
        """Get value from specific cell"""
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name]
        value = sheet[cell_address].value
        wb.close()
        return value
    
    def extract_hidden_rows(self, file_path: Path) -> Dict[str, List[int]]:
        """Extract information about hidden rows per sheet"""
        hidden_rows = {}
        
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            hidden = []
            
            for row_dim in sheet.row_dimensions:
                if sheet.row_dimensions[row_dim].hidden:
                    hidden.append(row_dim)
            
            if hidden:
                hidden_rows[sheet_name] = hidden
        
        wb.close()
        return hidden_rows
    
    def extract_hidden_columns(self, file_path: Path) -> Dict[str, List[str]]:
        """Extract information about hidden columns per sheet"""
        hidden_cols = {}
        
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            hidden = []
            
            for col_dim in sheet.column_dimensions:
                if sheet.column_dimensions[col_dim].hidden:
                    hidden.append(col_dim)
            
            if hidden:
                hidden_cols[sheet_name] = hidden
        
        wb.close()
        return hidden_cols


# Register parser
DocumentParserFactory.register(["xlsx", "xls", "xlsm"], ExcelParser)