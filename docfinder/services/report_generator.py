"""Report generation module."""
from typing import Dict, Any, List
import io
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    pd = None
    Workbook = None


class ReportGenerator:
    """Generate PDF and Excel reports."""

    @staticmethod
    def generate_pdf_report(comparison_result: Dict[str, Any], user_info: Dict[str, Any]) -> bytes:
        """
        Generate a PDF report from comparison results.
        
        Args:
            comparison_result: Comparison results dictionary
            user_info: User information
        
        Returns:
            PDF file as bytes
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2C3E50')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#3498DB')
        )
        
        normal_style = styles['Normal']
        
        story = []
        
        # Title
        story.append(Paragraph("doc-differ-ai Comparison Report", title_style))
        story.append(Spacer(1, 20))
        
        # Report Info
        story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        story.append(Paragraph(f"<b>User:</b> {user_info.get('username', 'N/A')}", normal_style))
        story.append(Paragraph(f"<b>Report Type:</b> {comparison_result.get('type', 'Unknown').upper()} Comparison", normal_style))
        story.append(Spacer(1, 30))
        
        # Summary Section
        story.append(Paragraph("Summary", heading_style))
        
        similarity = comparison_result.get("similarity_score", 0)
        summary_data = [
            ["Metric", "Value"],
            ["Similarity Score", f"{similarity * 100:.2f}%"],
            ["Total Additions", str(comparison_result.get("total_additions", 0))],
            ["Total Deletions", str(comparison_result.get("total_deletions", 0))],
            ["Type", comparison_result.get("type", "N/A").upper()]
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ECF0F1')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Additions
        additions = comparison_result.get("additions", [])
        if additions:
            story.append(Paragraph("Additions", heading_style))
            story.append(Paragraph(f"<font color='green'>+ {len(additions)} items added</font>", normal_style))
            for i, add in enumerate(additions[:10]):
                if isinstance(add, dict):
                    text = add.get("word", str(add))
                else:
                    text = str(add)
                story.append(Paragraph(f"• {text[:100]}", normal_style))
            if len(additions) > 10:
                story.append(Paragraph(f"... and {len(additions) - 10} more", normal_style))
            story.append(Spacer(1, 10))
        
        # Deletions
        deletions = comparison_result.get("deletions", [])
        if deletions:
            story.append(Paragraph("Deletions", heading_style))
            story.append(Paragraph(f"<font color='red'>- {len(deletions)} items deleted</font>", normal_style))
            for i, dele in enumerate(deletions[:10]):
                if isinstance(dele, dict):
                    text = dele.get("word", str(dele))
                else:
                    text = str(dele)
                story.append(Paragraph(f"• {text[:100]}", normal_style))
            if len(deletions) > 10:
                story.append(Paragraph(f"... and {len(deletions) - 10} more", normal_style))
        
        # Footer
        story.append(Spacer(1, 40))
        story.append(Paragraph("<i>Generated by doc-differ-ai - Intelligent Document Comparison Platform</i>", normal_style))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_excel_report(comparison_result: Dict[str, Any], user_info: Dict[str, Any]) -> bytes:
        """
        Generate an Excel report from comparison results.
        
        Args:
            comparison_result: Comparison results dictionary
            user_info: User information
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        title_font = Font(bold=True, size=16)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary['A1'] = "doc-differ-ai Comparison Report"
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = "Generated:"
        ws_summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_summary['A4'] = "User:"
        ws_summary['B4'] = user_info.get('username', 'N/A')
        ws_summary['A5'] = "Report Type:"
        ws_summary['B5'] = f"{comparison_result.get('type', 'Unknown')} Comparison"
        
        ws_summary['A7'] = "Metric"
        ws_summary['B7'] = "Value"
        for col in ['A', 'B']:
            cell = ws_summary[f'{col}7']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        metrics = [
            ("Similarity Score", f"{comparison_result.get('similarity_score', 0) * 100:.2f}%"),
            ("Total Additions", comparison_result.get("total_additions", 0)),
            ("Total Deletions", comparison_result.get("total_deletions", 0)),
            ("Type", comparison_result.get("type", "N/A").upper())
        ]
        
        for i, (metric, value) in enumerate(metrics, start=8):
            ws_summary[f'A{i}'] = metric
            ws_summary[f'B{i}'] = value
            ws_summary[f'A{i}'].border = thin_border
            ws_summary[f'B{i}'].border = thin_border
        
        # Additions Sheet
        ws_add = wb.create_sheet("Additions")
        ws_add['A1'] = "Additions"
        ws_add['A1'].font = title_font
        
        additions = comparison_result.get("additions", [])
        ws_add['A3'] = "#"
        ws_add['B3'] = "Content"
        for col in ['A', 'B']:
            cell = ws_add[f'{col}3']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for i, add in enumerate(additions[:100], start=4):
            ws_add[f'A{i}'] = i - 3
            if isinstance(add, dict):
                ws_add[f'B{i}'] = add.get("word", str(add))
            else:
                ws_add[f'B{i}'] = str(add)
            ws_add[f'A{i}'].border = thin_border
            ws_add[f'B{i}'].border = thin_border
        
        # Deletions Sheet
        ws_del = wb.create_sheet("Deletions")
        ws_del['A1'] = "Deletions"
        ws_del['A1'].font = title_font
        
        deletions = comparison_result.get("deletions", [])
        ws_del['A3'] = "#"
        ws_del['B3'] = "Content"
        for col in ['A', 'B']:
            cell = ws_del[f'{col}3']
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for i, dele in enumerate(deletions[:100], start=4):
            ws_del[f'A{i}'] = i - 3
            if isinstance(dele, dict):
                ws_del[f'B{i}'] = dele.get("word", str(dele))
            else:
                ws_del[f'B{i}'] = str(dele)
            ws_del[f'A{i}'].border = thin_border
            ws_del[f'B{i}'].border = thin_border
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 30
        ws_add.column_dimensions['A'].width = 10
        ws_add.column_dimensions['B'].width = 50
        ws_del.column_dimensions['A'].width = 10
        ws_del.column_dimensions['B'].width = 50
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()